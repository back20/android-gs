import os
import json
import pandas as pd
from datetime import datetime
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.metrics import dp
from kivy.properties import ObjectProperty, StringProperty, NumericProperty, ListProperty, BooleanProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.app import MDApp
from kivymd.uix.button import MDRectangleFlatButton, MDFlatButton, MDIconButton
from kivymd.uix.card import MDCard
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.dialog import MDDialog
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.label import MDLabel
from kivymd.uix.picker import MDDatePicker, MDTimePicker
from kivymd.uix.snackbar import MDSnackbar
from kivymd.uix.textfield import MDTextField
from kivymd.uix.toolbar import MDToolbar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 确保中文显示正常
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
plt.rcParams["font.family"] = ["SimHei", "WenQuanYi Micro Hei", "Heiti TC"]

# 为移动设备优化布局
Window.size = (360, 640)  # 模拟手机尺寸

class TimeRecordRow(BoxLayout):
    """工时记录行组件"""
    date_text = StringProperty("选择日期")
    plate_name_text = StringProperty("")
    start_time_text = StringProperty("开始时间")
    end_time_text = StringProperty("结束时间")
    overtime_text = StringProperty("0")
    rate_text = StringProperty("0")
    amount_text = StringProperty("0")
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.date_dialog = None
        self.start_time_dialog = None
        self.end_time_dialog = None
        self.parent_screen = None  # 存储对父屏幕的引用
        
    def set_parent_screen(self, screen):
        """设置父屏幕引用"""
        self.parent_screen = screen
        
    def show_date_picker(self):
        """显示日期选择器"""
        if not self.date_dialog:
            self.date_dialog = MDDatePicker(
                callback=self.get_date,
                year=datetime.now().year,
                month=datetime.now().month,
                day=datetime.now().day
            )
        self.date_dialog.open()
        
    def get_date(self, date_obj):
        """处理选择的日期"""
        self.date_text = str(date_obj)
        
    def show_start_time_picker(self):
        """显示开始时间选择器"""
        if not self.start_time_dialog:
            self.start_time_dialog = MDTimePicker()
            self.start_time_dialog.bind(time=self.get_start_time)
        self.start_time_dialog.open()
        
    def get_start_time(self, instance, time):
        """处理选择的开始时间"""
        self.start_time_text = str(time)
        
    def show_end_time_picker(self):
        """显示结束时间选择器"""
        if not self.end_time_dialog:
            self.end_time_dialog = MDTimePicker()
            self.end_time_dialog.bind(time=self.get_end_time)
        self.end_time_dialog.open()
        
    def get_end_time(self, instance, time):
        """处理选择的结束时间"""
        self.end_time_text = str(time)
        
    def calculate_time_and_amount(self):
        """计算工作时间和金额"""
        if self.start_time_text != "开始时间" and self.end_time_text != "结束时间":
            try:
                start_time = datetime.strptime(self.start_time_text, "%H:%M:%S")
                end_time = datetime.strptime(self.end_time_text, "%H:%M:%S")
                work_hours = (end_time - start_time).total_seconds() / 3600
                
                # 计算总时间（包括加班）
                try:
                    overtime = float(self.overtime_text)
                except ValueError:
                    overtime = 0
                    
                total_hours = work_hours + overtime
                
                # 计算金额
                try:
                    rate = float(self.rate_text)
                    amount = total_hours * rate
                    self.amount_text = f"{amount:.2f}"
                    
                    # 更新父屏幕的总计
                    if self.parent_screen:
                        self.parent_screen.update_totals()
                except ValueError:
                    pass
            except ValueError:
                pass

class ProjectItem(BoxLayout):
    """项目项组件"""
    project_name = StringProperty("")
    start_date = StringProperty("开始日期")
    end_date = StringProperty("结束日期")
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.date_picker = None
        self.is_start_date = True  # 标记当前选择的是开始日期还是结束日期
        
    def show_date_picker(self, is_start=True):
        """显示日期选择器"""
        self.is_start_date = is_start
        if not self.date_picker:
            self.date_picker = MDDatePicker(
                callback=self.get_date,
                year=datetime.now().year,
                month=datetime.now().month,
                day=datetime.now().day
            )
        self.date_picker.open()
        
    def get_date(self, date_obj):
        """处理选择的日期"""
        if self.is_start_date:
            self.start_date = str(date_obj)
        else:
            self.end_date = str(date_obj)
            
    def edit_project(self):
        """编辑项目"""
        app = MDApp.get_running_app()
        app.root.current = "project"
        app.root.ids.project.load_project(self.project_name)

class RateSettingDialog(MDDialog):
    """单价设置对话框"""
    rate_name = StringProperty("")
    rate_value = StringProperty("0")
    
    def __init__(self, rate_data=None, callback=None, **kwargs):
        super().__init__(**kwargs)
        self.callback = callback
        if rate_data:
            self.rate_name = rate_data[0]
            self.rate_value = str(rate_data[1])
            
    def save_rate(self):
        """保存单价设置"""
        if self.rate_name and self.rate_value:
            try:
                rate_value = float(self.rate_value)
                if self.callback:
                    self.callback((self.rate_name, rate_value))
                self.dismiss()
            except ValueError:
                MDSnackbar(text="请输入有效的单价数值！").open()

class MainScreen(Screen):
    """主屏幕"""
    projects_layout = ObjectProperty(None)
    project_name_input = ObjectProperty(None)
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.projects = {}  # 存储所有项目
        self.load_projects()  # 加载保存的项目
        
    def load_projects(self):
        """加载保存的项目"""
        try:
            if os.path.exists("projects.json"):
                with open("projects.json", "r", encoding="utf-8") as f:
                    self.projects = json.load(f)
                    
                # 显示所有项目
                self.update_projects_list()
        except Exception as e:
            print(f"加载项目失败: {e}")
            
    def save_projects(self):
        """保存项目列表"""
        try:
            with open("projects.json", "w", encoding="utf-8") as f:
                json.dump(self.projects, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存项目失败: {e}")
            
    def add_project(self):
        """添加新项目"""
        project_name = self.project_name_input.text.strip()
        if not project_name:
            MDSnackbar(text="请输入项目名称！").open()
            return
            
        if project_name in self.projects:
            MDSnackbar(text="项目名称已存在！").open()
            return
            
        # 初始化项目数据
        self.projects[project_name] = {
            "start_date": "开始日期",
            "end_date": "结束日期",
            "records": []
        }
        
        # 保存项目
        self.save_projects()
        
        # 更新项目列表
        self.update_projects_list()
        
        # 清空输入框
        self.project_name_input.text = ""
        
    def update_projects_list(self):
        """更新项目列表显示"""
        self.projects_layout.clear_widgets()
        
        for project_name, project_data in self.projects.items():
            item = ProjectItem()
            item.project_name = project_name
            item.start_date = project_data["start_date"]
            item.end_date = project_data["end_date"]
            
            # 添加到布局
            self.projects_layout.add_widget(item)
            
    def export_all_projects(self):
        """导出所有项目"""
        if not self.projects:
            MDSnackbar(text="没有项目可导出！").open()
            return
            
        for project_name, project_data in self.projects.items():
            self.export_project(project_name, project_data)
            
        MDSnackbar(text="所有项目已导出！").open()
        
    def export_project(self, project_name, project_data):
        """导出单个项目"""
        records = project_data["records"]
        if not records:
            return
            
        # 准备数据
        data = []
        for record in records:
            if record['date'] != "选择日期" and record['plate_name']:
                try:
                    start_time = datetime.strptime(record['start_time'], "%H:%M:%S")
                    end_time = datetime.strptime(record['end_time'], "%H:%M:%S")
                    work_hours = (end_time - start_time).total_seconds() / 3600
                    
                    overtime = float(record['overtime'])
                    total_hours = work_hours + overtime
                    amount = float(record['amount'])
                    
                    data.append({
                        'date': record['date'],
                        'plate_name': record['plate_name'],
                        'work_hours': work_hours,
                        'overtime': overtime,
                        'total_hours': total_hours,
                        'rate': float(record['rate']),
                        'amount': amount
                    })
                except ValueError:
                    pass
        
        if not data:
            return
            
        # 按车牌/姓名分组统计
        grouped_data = {}
        for item in data:
            plate_name = item['plate_name']
            if plate_name not in grouped_data:
                grouped_data[plate_name] = {
                    'dates': set(),
                    'total_hours': 0,
                    'amount': 0,
                    'records': []
                }
                
            grouped_data[plate_name]['dates'].add(item['date'])
            grouped_data[plate_name]['total_hours'] += item['total_hours']
            grouped_data[plate_name]['amount'] += item['amount']
            grouped_data[plate_name]['records'].append(item)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工时统计"
        
        # 添加标题
        ws['A1'] = f"{project_name} 工时统计报表"
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加日期范围
        ws['A2'] = f"日期范围: {project_data['start_date']} - {project_data['end_date']}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加表头
        headers = ["日期", "车牌/姓名", "工作时间", "加班时间", "总时间", "单价", "金额"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加数据
        row_idx = 5
        for plate_name, info in grouped_data.items():
            # 添加该车牌/姓名的所有记录
            for record in info['records']:
                ws[f'A{row_idx}'] = record['date']
                ws[f'B{row_idx}'] = record['plate_name']
                ws[f'C{row_idx}'] = round(record['work_hours'], 2)
                ws[f'D{row_idx}'] = round(record['overtime'], 2)
                ws[f'E{row_idx}'] = round(record['total_hours'], 2)
                ws[f'F{row_idx}'] = round(record['rate'], 2)
                ws[f'G{row_idx}'] = round(record['amount'], 2)
                row_idx += 1
            
            # 添加分隔行
            ws[f'A{row_idx}'] = "小计"
            ws[f'B{row_idx}'] = plate_name
            ws[f'C{row_idx}'] = ""
            ws[f'D{row_idx}'] = ""
            ws[f'E{row_idx}'] = round(info['total_hours'], 2)
            ws[f'F{row_idx}'] = ""
            ws[f'G{row_idx}'] = round(info['amount'], 2)
            
            # 设置小计行样式
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            row_idx += 2  # 空一行
            
        # 添加总计行
        row_idx += 1
        ws[f'A{row_idx}'] = "总计"
        ws[f'B{row_idx}'] = f"{len(grouped_data)}人"
        ws[f'C{row_idx}'] = ""
        ws[f'D{row_idx}'] = ""
        
        total_hours = sum(info['total_hours'] for info in grouped_data.values())
        total_amount = sum(info['amount'] for info in grouped_data.values())
        
        ws[f'E{row_idx}'] = round(total_hours, 2)
        ws[f'F{row_idx}'] = ""
        ws[f'G{row_idx}'] = round(total_amount, 2)
        
        # 设置总计行样式
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动调整列宽
        for col_idx in range(1, 8):
            max_length = 0
            column = get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                if len(str(ws[f'{column}{row}'].value)) > max_length:
                    max_length = len(str(ws[f'{column}{row}'].value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 25)  # 限制最大宽度
        
        # 保存Excel文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{project_name}_{timestamp}.xlsx"
        wb.save(filename)

class ProjectScreen(Screen):
    """项目详情屏幕"""
    project_name = StringProperty("")
    start_date = StringProperty("开始日期")
    end_date = StringProperty("结束日期")
    time_records_layout = ObjectProperty(None)
    total_days_text = StringProperty("0")
    total_hours_text = StringProperty("0")
    total_amount_text = StringProperty("0")
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.time_records = []  # 存储所有工时记录
        self.rate_settings = {}  # 存储单价设置
        self.load_rate_settings()  # 加载保存的单价设置
        self.file_manager = None
        self.dialog = None
        self.preview_dialog = None
        
    def load_project(self, project_name):
        """加载项目数据"""
        self.project_name = project_name
        
        # 从主屏幕获取项目数据
        main_screen = self.manager.get_screen("main")
        project_data = main_screen.projects.get(project_name, {})
        
        self.start_date = project_data.get("start_date", "开始日期")
        self.end_date = project_data.get("end_date", "结束日期")
        
        # 清空现有记录
        self.time_records_layout.clear_widgets()
        self.time_records = []
        
        # 加载项目记录
        for record_data in project_data.get("records", []):
            row = TimeRecordRow()
            row.set_parent_screen(self)
            row.date_text = record_data.get("date", "选择日期")
            row.plate_name_text = record_data.get("plate_name", "")
            row.start_time_text = record_data.get("start_time", "开始时间")
            row.end_time_text = record_data.get("end_time", "结束时间")
            row.overtime_text = record_data.get("overtime", "0")
            row.rate_text = record_data.get("rate", "0")
            row.amount_text = record_data.get("amount", "0")
            
            self.time_records_layout.add_widget(row)
            self.time_records.append(row)
            
        # 更新总计
        self.update_totals()
        
    def load_rate_settings(self):
        """加载保存的单价设置"""
        try:
            if os.path.exists("rate_settings.json"):
                with open("rate_settings.json", "r", encoding="utf-8") as f:
                    self.rate_settings = json.load(f)
        except Exception as e:
            print(f"加载单价设置失败: {e}")
            
    def save_rate_settings(self):
        """保存单价设置"""
        try:
            with open("rate_settings.json", "w", encoding="utf-8") as f:
                json.dump(self.rate_settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存单价设置失败: {e}")
            
    def save_project(self):
        """保存项目数据"""
        main_screen = self.manager.get_screen("main")
        
        # 准备项目数据
        project_data = {
            "start_date": self.start_date,
            "end_date": self.end_date,
            "records": []
        }
        
        # 收集所有记录
        for record in self.time_records:
            record_data = {
                "date": record.date_text,
                "plate_name": record.plate_name_text,
                "start_time": record.start_time_text,
                "end_time": record.end_time_text,
                "overtime": record.overtime_text,
                "rate": record.rate_text,
                "amount": record.amount_text
            }
            project_data["records"].append(record_data)
            
        # 保存到主屏幕
        main_screen.projects[self.project_name] = project_data
        main_screen.save_projects()
        
    def add_time_record_row(self):
        """添加新的工时记录行"""
        row = TimeRecordRow()
        row.set_parent_screen(self)
        self.time_records_layout.add_widget(row)
        self.time_records.append(row)
        return row
        
    def add_new_record(self):
        """添加新记录按钮回调"""
        self.add_time_record_row()
        
    def show_rate_settings(self):
        """显示单价设置对话框"""
        rate_items = [
            MDFlatButton(
                text=f"{name}: {rate}",
                on_press=lambda x, n=name, r=rate: self.edit_rate(n, r)
            ) for name, rate in self.rate_settings.items()
        ]
        
        rate_items.append(
            MDFlatButton(
                text="新增单价",
                on_press=lambda x: self.create_new_rate()
            )
        )
        
        if self.dialog:
            self.dialog.dismiss()
            
        self.dialog = MDDialog(
            title="单价设置",
            type="simple",
            items=rate_items,
            buttons=[
                MDFlatButton(
                    text="关闭",
                    on_press=lambda x: self.dialog.dismiss()
                )
            ]
        )
        self.dialog.open()
        
    def create_new_rate(self):
        """创建新单价"""
        if self.dialog:
            self.dialog.dismiss()
            
        rate_dialog = RateSettingDialog(
            callback=self.save_new_rate,
            size_hint=(0.8, 0.4)
        )
        rate_dialog.open()
        
    def save_new_rate(self, rate_data):
        """保存新单价"""
        name, rate = rate_data
        self.rate_settings[name] = rate
        self.save_rate_settings()
        
    def edit_rate(self, name, rate):
        """编辑现有单价"""
        if self.dialog:
            self.dialog.dismiss()
            
        rate_dialog = RateSettingDialog(
            rate_data=(name, rate),
            callback=lambda data: self.update_rate(name, data),
            size_hint=(0.8, 0.4)
        )
        rate_dialog.open()
        
    def update_rate(self, old_name, rate_data):
        """更新现有单价"""
        new_name, rate = rate_data
        if old_name != new_name:
            del self.rate_settings[old_name]
        self.rate_settings[new_name] = rate
        self.save_rate_settings()
        
    def apply_rate(self, name, rate):
        """应用单价到当前选中的记录"""
        if self.dialog:
            self.dialog.dismiss()
            
        # 应用到最后一个记录行
        if self.time_records:
            self.time_records[-1].rate_text = str(rate)
            self.time_records[-1].calculate_time_and_amount()
            
    def update_totals(self):
        """更新总计信息"""
        total_days = 0
        total_hours = 0
        total_amount = 0
        
        unique_dates = set()
        
        for record in self.time_records:
            if record.date_text != "选择日期":
                unique_dates.add(record.date_text)
                
            if record.start_time_text != "开始时间" and record.end_time_text != "结束时间":
                try:
                    start_time = datetime.strptime(record.start_time_text, "%H:%M:%S")
                    end_time = datetime.strptime(record.end_time_text, "%H:%M:%S")
                    work_hours = (end_time - start_time).total_seconds() / 3600
                    
                    try:
                        overtime = float(record.overtime_text)
                    except ValueError:
                        overtime = 0
                        
                    total_hours += work_hours + overtime
                except ValueError:
                    pass
                    
            try:
                total_amount += float(record.amount_text)
            except ValueError:
                pass
                
        total_days = len(unique_dates)
        
        self.total_days_text = str(total_days)
        self.total_hours_text = f"{total_hours:.2f}"
        self.total_amount_text = f"{total_amount:.2f}"
        
    def preview_excel(self):
        """预览Excel内容"""
        if not self.time_records:
            MDSnackbar(text="没有记录可预览！").open()
            return
            
        # 保存项目数据
        self.save_project()
            
        # 准备数据
        data = []
        for record in self.time_records:
            if record.date_text != "选择日期" and record.plate_name_text:
                try:
                    start_time = datetime.strptime(record.start_time_text, "%H:%M:%S")
                    end_time = datetime.strptime(record.end_time_text, "%H:%M:%S")
                    work_hours = (end_time - start_time).total_seconds() / 3600
                    
                    try:
                        overtime = float(record.overtime_text)
                    except ValueError:
                        overtime = 0
                        
                    total_hours = work_hours + overtime
                    amount = float(record.amount_text)
                    
                    data.append([
                        record.date_text,
                        record.plate_name_text,
                        f"{work_hours:.2f}",
                        f"{overtime:.2f}",
                        f"{total_hours:.2f}",
                        record.rate_text,
                        f"{amount:.2f}"
                    ])
                except ValueError:
                    pass
        
        if not data:
            MDSnackbar(text="没有有效的记录可预览！").open()
            return
            
        # 创建数据表格
        table = MDDataTable(
            size_hint=(0.9, 0.8),
            use_pagination=True,
            column_data=[
                ("日期", dp(20)),
                ("车牌/姓名", dp(20)),
                ("工作时间", dp(15)),
                ("加班时间", dp(15)),
                ("总时间", dp(15)),
                ("单价", dp(15)),
                ("金额", dp(15))
            ],
            row_data=data
        )
        
        if self.preview_dialog:
            self.preview_dialog.dismiss()
            
        self.preview_dialog = MDDialog(
            title="Excel预览",
            type="custom",
            content_cls=table,
            buttons=[
                MDFlatButton(
                    text="关闭",
                    on_press=lambda x: self.preview_dialog.dismiss()
                ),
                MDFlatButton(
                    text="生成Excel",
                    on_press=lambda x: [self.preview_dialog.dismiss(), self.generate_excel()]
                )
            ]
        )
        self.preview_dialog.open()
        
    def generate_excel(self):
        """生成Excel文件"""
        # 保存项目数据
        self.save_project()
            
        # 获取主屏幕项目数据
        main_screen = self.manager.get_screen("main")
        project_data = main_screen.projects.get(self.project_name, {})
        records = project_data.get("records", [])
        
        if not records:
            MDSnackbar(text="没有记录可生成Excel！").open()
            return
            
        # 准备数据
        data = []
        for record in records:
            if record['date'] != "选择日期" and record['plate_name']:
                try:
                    start_time = datetime.strptime(record['start_time'], "%H:%M:%S")
                    end_time = datetime.strptime(record['end_time'], "%H:%M:%S")
                    work_hours = (end_time - start_time).total_seconds() / 3600
                    
                    overtime = float(record['overtime'])
                    total_hours = work_hours + overtime
                    amount = float(record['amount'])
                    
                    data.append({
                        'date': record['date'],
                        'plate_name': record['plate_name'],
                        'work_hours': work_hours,
                        'overtime': overtime,
                        'total_hours': total_hours,
                        'rate': float(record['rate']),
                        'amount': amount
                    })
                except ValueError:
                    pass
        
        if not data:
            MDSnackbar(text="没有有效的记录可生成Excel！").open()
            return
            
        # 按车牌/姓名分组统计
        grouped_data = {}
        for item in data:
            plate_name = item['plate_name']
            if plate_name not in grouped_data:
                grouped_data[plate_name] = {
                    'dates': set(),
                    'total_hours': 0,
                    'amount': 0,
                    'records': []
                }
                
            grouped_data[plate_name]['dates'].add(item['date'])
            grouped_data[plate_name]['total_hours'] += item['total_hours']
            grouped_data[plate_name]['amount'] += item['amount']
            grouped_data[plate_name]['records'].append(item)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工时统计"
        
        # 添加标题
        ws['A1'] = f"{self.project_name} 工时统计报表"
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加日期范围
        ws['A2'] = f"日期范围: {self.start_date} - {self.end_date}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加表头
        headers = ["日期", "车牌/姓名", "工作时间", "加班时间", "总时间", "单价", "金额"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加数据
        row_idx = 5
        for plate_name, info in grouped_data.items():
            # 添加该车牌/姓名的所有记录
            for record in info['records']:
                ws[f'A{row_idx}'] = record['date']
                ws[f'B{row_idx}'] = record['plate_name']
                ws[f'C{row_idx}'] = round(record['work_hours'], 2)
                ws[f'D{row_idx}'] = round(record['overtime'], 2)
                ws[f'E{row_idx}'] = round(record['total_hours'], 2)
                ws[f'F{row_idx}'] = round(record['rate'], 2)
                ws[f'G{row_idx}'] = round(record['amount'], 2)
                row_idx += 1
            
            # 添加分隔行
            ws[f'A{row_idx}'] = "小计"
            ws[f'B{row_idx}'] = plate_name
            ws[f'C{row_idx}'] = ""
            ws[f'D{row_idx}'] = ""
            ws[f'E{row_idx}'] = round(info['total_hours'], 2)
            ws[f'F{row_idx}'] = ""
            ws[f'G{row_idx}'] = round(info['amount'], 2)
            
            # 设置小计行样式
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            row_idx += 2  # 空一行
            
        # 添加总计行
        row_idx += 1
        ws[f'A{row_idx}'] = "总计"
        ws[f'B{row_idx}'] = f"{len(grouped_data)}人"
        ws[f'C{row_idx}'] = ""
        ws[f'D{row_idx}'] = ""
        
        total_hours = sum(info['total_hours'] for info in grouped_data.values())
        total_amount = sum(info['amount'] for info in grouped_data.values())
        
        ws[f'E{row_idx}'] = round(total_hours, 2)
        ws[f'F{row_idx}'] = ""
        ws[f'G{row_idx}'] = round(total_amount, 2)
        
        # 设置总计行样式
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动调整列宽
        for col_idx in range(1, 8):
            max_length = 0
            column = get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                if len(str(ws[f'{column}{row}'].value)) > max_length:
                    max_length = len(str(ws[f'{column}{row}'].value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 25)  # 限制最大宽度
        
        # 保存Excel文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.project_name}_{timestamp}.xlsx"
        wb.save(filename)
        
        MDSnackbar(text=f"Excel文件已生成: {filename}").open()

class TimeSheetApp(MDApp):
    """工时统计应用主类"""
    def build(self):
        """构建应用UI"""
        self.title = "工时价格Excel生成器"
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.primary_hue = "500"
        self.theme_cls.theme_style = "Light"
        
        # 创建屏幕管理器
        sm = ScreenManager()
        sm.add_widget(MainScreen(name="main"))
        sm.add_widget(ProjectScreen(name="project"))
        
        return sm
        
    def on_start(self):
        """应用启动时调用"""
        # 为移动设备优化布局
        self.root.ids.main.ids.projects_layout.spacing = dp(10)
        self.root.ids.project.ids.time_records_layout.spacing = dp(10)

if __name__ == "__main__":
    TimeSheetApp().run()
